# -*- coding:utf-8 –*-
import os
from openpyxl import Workbook
import xlrd

tools_type = ["fio"]
fs_type = ["ext4", "ext4nj"]
iozone_rw = "rw"
rw_mode = {"fio": ["write"]}


class Analysis:
    def __init__(self, result_directory, scale_test=True):
        self._root_dir = result_directory
        self._scale_test = scale_test
        self._max_num = 16
        self._tools_type = tools_type
        self._fs_type = fs_type
        self._rw_mode = rw_mode

    def _read_per_file(self, file, tool_type):
        # iops  k
        # BW  MiB/s
        global iops, bw

        if tool_type == "fio":
            fp = open(file, "r")
            while True:
                line = fp.readline()
                if "IOPS" not in line:
                    continue
                iops_bw = line.split(":")[1].split("(")[0].split(",")
                iops_bw[0] = iops_bw[0].replace("IOPS=", "")
                iops_bw[1] = iops_bw[1].replace("BW=", "")
                # print(iops_bw[0], iops_bw[1])
                if "k" in iops_bw[0]:
                    iops = round(float(iops_bw[0].replace('k', '')), 3)
                else:
                    iops = round(float(iops_bw[0]) / 1000, 3)
                if "MiB/s" in iops_bw[1]:
                    bw = round(float(iops_bw[1].replace("MiB/s", "")), 3)
                elif "KiB/s" in iops_bw[1]:
                    bw = round(float(iops_bw[1].replace("KiB/s", "")) / 1024, 3)
                break
            fp.close()
            return iops, bw
        elif tool_type == "sysbench":
            fp = open(file, "r")
            while True:
                line = fp.readline()
                if "File operations:" in line:
                    line = fp.readline()
                    if "0.00" in line.split(":")[1]:
                        line = fp.readline()
                        iops = round(float(line.split(":")[1]) / 1000, 3)
                    else:
                        iops = round(float(line.split(":")[1]) / 1000, 3)
                elif "Throughput:" in line:
                    line = fp.readline()
                    if "0.00" in line.split(":")[1]:
                        line = fp.readline()
                        bw = round(float(line.split(":")[1]) / 1024, 3)
                    else:
                        bw = round(float(line.split(":")[1]) / 1024, 3)
                    break
                else:
                    continue
            fp.close()
            return iops, bw
        elif tool_type == "iozone":
            wb = xlrd.open_workbook(filename=file, encoding_override="utf-8")
            table = wb.sheet_by_index(0)
            iops = float(table.cell(11, 1).value) / 1024
            bw = float(table.cell(5, 1).value) / 1024
            return iops, bw

    def _read_file_list(self, fs, rw, num, tool_type, file_list):
        sum_iops = 0
        sum_bw = 0
        # print(fs, rw, num)
        for file in file_list:
            if "%s-%s-%s-" % (fs, rw, str(num)) in file:
                iops, bw = self._read_per_file(os.path.join(self._root_dir, tool_type, file), "fio")
                # print(file, iops, bw)
                sum_iops += iops
                sum_bw += bw
        avg_iops = "%.3f" % (sum_iops / num)
        avg_bw = "%.3f" % (sum_bw / num)
        print(avg_iops, avg_bw)
        return avg_iops, avg_bw

    def _write2file(self, data, file):
        wb = Workbook()
        bw_ws = wb.get_sheet_by_name("Sheet")
        bw_ws.title = "bw"
        iops_ws = wb.create_sheet("iops")
        for i in range(1, self._max_num + 1):
            bw_ws.cell(row=1, column=i + 1).value = i
            iops_ws.cell(row=1, column=i + 1).value = i

        for line in data:
            row = 2 + 2 * self._fs_type.index(line[0]) + self._rw_mode["fio"].index(line[1])
            col = int(line[2]) + 1
            bw_ws.cell(row=row, column=1).value = "%s-%s" % (line[0], line[1])
            iops_ws.cell(row=row, column=1).value = "%s-%s" % (line[0], line[1])
            bw_ws.cell(row=row, column=col).value = line[4]
            iops_ws.cell(row=row, column=col).value = line[3]
        wb.save(file)

    def start(self):
        dir_list = os.listdir(self._root_dir)
        for tool_type in dir_list:
            print(tool_type)
            line_list = []
            tool_dir = os.path.join(self._root_dir, tool_type)
            if not os.path.isdir(tool_dir):
                continue
            file_list = os.listdir(tool_dir)
            file_list.sort()
            for file in file_list:
                if "-lockstat" in file:
                    file_list.remove(file)
            for fs in self._fs_type:
                for rw in ["write"]:
                    if not self._scale_test:
                        tp = self._read_file_list(fs, rw, self._max_num, tool_type, file_list)
                        line_list.append([fs, rw, str(self._max_num), tp[0], tp[1]])
                    else:
                        for i in range(0, self._max_num):
                            tp = self._read_file_list(fs, rw, i + 1, tool_type, file_list)
                            line_list.append([fs, rw, str(i+1), tp[0], tp[1]])
            self._write2file(line_list, "%s/%s.xlsx" % (self._root_dir, tool_type))
